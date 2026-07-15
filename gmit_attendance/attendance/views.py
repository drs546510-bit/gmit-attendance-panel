import json
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import (
    TeacherSignupForm, StudentUserSignupForm, ExcelImportForm,
    TeacherAssignmentForm, SubjectForm,
)
from .models import (
    CustomUser, TeacherProfile, TeacherAssignment, Student, Attendance, Subject,
    BRANCH_CHOICES, SEMESTER_CHOICES,
)

BRANCH_DICT = dict(BRANCH_CHOICES)


def is_admin(user):
    # Superusers (created via `createsuperuser`) are always treated as Admins,
    # even before anyone manually sets role='ADMIN' on them.
    return user.is_authenticated and (user.role == 'ADMIN' or user.is_superuser)


class GmitLoginView(LoginView):
    template_name = 'attendance/login.html'


@login_required
def dashboard_redirect(request):
    user = request.user
    if user.role == 'ADMIN' or user.is_superuser:
        return redirect('admin_dashboard')
    if user.role == 'TEACHER':
        return redirect('teacher_dashboard')
    return redirect('student_dashboard')


def signup_teacher(request):
    if request.method == 'POST':
        form = TeacherSignupForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'TEACHER'
            user.email = form.cleaned_data['email']
            user.phone = form.cleaned_data.get('phone', '')
            user.save()
            TeacherProfile.objects.create(user=user, is_approved=False)
            messages.success(
                request,
                "Account created! Your teacher account needs Admin approval "
                "before you can mark attendance."
            )
            return redirect('login')
    else:
        form = TeacherSignupForm()
    return render(request, 'attendance/signup_teacher.html', {'form': form})


def signup_student(request):
    if request.method == 'POST':
        form = StudentUserSignupForm(request.POST)
        if form.is_valid():
            usn = form.cleaned_data['usn'].strip().upper()
            try:
                student = Student.objects.get(usn=usn)
            except Student.DoesNotExist:
                form.add_error('usn', 'No student record found with this USN. Ask your teacher/admin to add you first.')
            else:
                if student.user_account_id:
                    form.add_error('usn', 'This USN is already linked to another login account.')
                else:
                    user = form.save(commit=False)
                    user.role = 'STUDENT'
                    user.email = form.cleaned_data['email']
                    user.save()
                    student.user_account = user
                    student.save()
                    messages.success(request, "Account created! You can now log in to view your attendance.")
                    return redirect('login')
    else:
        form = StudentUserSignupForm()
    return render(request, 'attendance/signup_student.html', {'form': form})


# ---------------- ADMIN ----------------

@login_required
def admin_dashboard(request):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")

    pending_teachers = TeacherProfile.objects.filter(is_approved=False).select_related('user')
    approved_teachers = TeacherProfile.objects.filter(is_approved=True).select_related('user')

    branch_stats = []
    for code, label in BRANCH_CHOICES:
        total_students = Student.objects.filter(branch=code).count()
        subject_count = Subject.objects.filter(branch=code).count()
        branch_stats.append({
            'code': code, 'label': label,
            'total_students': total_students, 'subject_count': subject_count,
        })

    context = {
        'pending_teachers': pending_teachers,
        'approved_teachers': approved_teachers,
        'branch_stats': branch_stats,
        'total_students': Student.objects.count(),
        'total_teachers': approved_teachers.count(),
        'total_subjects': Subject.objects.count(),
        'today_marked': Attendance.objects.filter(date=timezone.localdate()).count(),
        'branch_choices': BRANCH_CHOICES,
        'semester_choices': SEMESTER_CHOICES,
    }
    return render(request, 'attendance/admin_dashboard.html', context)


@login_required
def approve_teacher(request, profile_id):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    profile = get_object_or_404(TeacherProfile, id=profile_id)
    profile.is_approved = True
    profile.approved_on = timezone.now()
    profile.save()
    messages.success(request, f"{profile.user.username} has been approved as a teacher.")
    return redirect('admin_dashboard')


@login_required
def reject_teacher(request, profile_id):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    profile = get_object_or_404(TeacherProfile, id=profile_id)
    username = profile.user.username
    profile.user.delete()
    messages.warning(request, f"Teacher request for {username} was rejected and removed.")
    return redirect('admin_dashboard')


@login_required
def manage_subjects(request):
    """Admin creates subjects per branch+semester, e.g. EE Sem1 -> Physics, Maths."""
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject added.")
            return redirect('manage_subjects')
    else:
        form = SubjectForm()

    subjects = Subject.objects.all()
    return render(request, 'attendance/manage_subjects.html', {
        'form': form, 'subjects': subjects,
    })


@login_required
def delete_subject(request, subject_id):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    subject = get_object_or_404(Subject, id=subject_id)
    subject.delete()
    messages.warning(request, "Subject deleted (any related assignments/attendance were removed too).")
    return redirect('manage_subjects')


@login_required
def export_reports(request):
    """A page listing every subject with a one-click Excel export button."""
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    subjects = Subject.objects.all()
    return render(request, 'attendance/export_reports.html', {'subjects': subjects})


@login_required
def export_subject_attendance(request, subject_id):
    """Builds and downloads an Excel file for one subject: every student as
    a row, every date+session that was ever marked as a column (Present/
    Absent), plus Total/Present/Absent/Percentage summary columns."""
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")

    subject = get_object_or_404(Subject, id=subject_id)
    students = Student.objects.filter(branch=subject.branch, semester=subject.semester).order_by('usn')
    records = Attendance.objects.filter(subject=subject).select_related('student').order_by(
        'date', 'session_number'
    )

    # Build an ordered, de-duplicated list of (date, session_number) columns
    # — one column per class session that was ever marked for this subject.
    seen = []
    seen_set = set()
    for r in records:
        key = (r.date, r.session_number)
        if key not in seen_set:
            seen_set.add(key)
            seen.append(key)

    # Quick lookup: (student_id, date, session_number) -> status
    status_lookup = {(r.student_id, r.date, r.session_number): r.status for r in records}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = subject.name[:31] or "Attendance"  # Excel sheet names cap at 31 chars

    header_fill = PatternFill(start_color="4F5EFF", end_color="4F5EFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    present_fill = PatternFill(start_color="E6FBF3", end_color="E6FBF3", fill_type="solid")
    absent_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    # ---- Title rows ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 + len(seen))
    title_cell = ws.cell(row=1, column=1, value=f"GMIT Student Attendance Panel — {subject.name} ({subject.branch}, Semester {subject.semester})")
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6 + len(seen))
    ws.cell(row=2, column=1, value=f"Exported on {timezone.localtime().strftime('%d-%b-%Y %I:%M %p')}").font = Font(italic=True, size=9, color="666666")

    header_row = 4
    headers = ["Sl.No", "USN", "Student Name"]
    for d, session in seen:
        label = d.strftime('%d-%b')
        headers.append(f"{label}\n(S{session})" if session > 1 else label)
    headers += ["Total Classes", "Present", "Absent", "Percentage"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, student in enumerate(students, start=header_row + 1):
        col = 1
        ws.cell(row=row_idx, column=col, value=row_idx - header_row).alignment = center; col += 1
        ws.cell(row=row_idx, column=col, value=student.usn).alignment = center; col += 1
        ws.cell(row=row_idx, column=col, value=student.name); col += 1

        present_count, absent_count = 0, 0
        for d, session in seen:
            status = status_lookup.get((student.id, d, session))
            cell = ws.cell(row=row_idx, column=col)
            if status == 'PRESENT':
                cell.value = 'P'
                cell.fill = present_fill
                present_count += 1
            elif status == 'ABSENT':
                cell.value = 'A'
                cell.fill = absent_fill
                absent_count += 1
            else:
                cell.value = '-'
            cell.alignment = center
            cell.border = thin_border
            col += 1

        total = present_count + absent_count
        pct = round((present_count / total) * 100, 1) if total else 0
        ws.cell(row=row_idx, column=col, value=total).alignment = center; col += 1
        ws.cell(row=row_idx, column=col, value=present_count).alignment = center; col += 1
        ws.cell(row=row_idx, column=col, value=absent_count).alignment = center; col += 1
        pct_cell = ws.cell(row=row_idx, column=col, value=f"{pct}%")
        pct_cell.alignment = center
        pct_cell.font = Font(bold=True, color="16B978" if pct >= 75 else "EF4444")

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    for i in range(len(seen)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 10
    for i in range(4):
        ws.column_dimensions[get_column_letter(4 + len(seen) + i)].width = 13

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = subject.name.replace(' ', '_')
    filename = f"{subject.branch}_Sem{subject.semester}_{safe_name}_attendance.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def assign_teacher(request, profile_id):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    profile = get_object_or_404(TeacherProfile, id=profile_id, is_approved=True)

    # Branch/semester picked via GET so the subject dropdown can be filtered
    # server-side, with a simple page reload — no JS required.
    selected_branch = request.POST.get('branch') or request.GET.get('branch')
    selected_semester = request.POST.get('semester') or request.GET.get('semester')

    if request.method == 'POST':
        form = TeacherAssignmentForm(
            request.POST, selected_branch=selected_branch, selected_semester=selected_semester
        )
        if form.is_valid():
            TeacherAssignment.objects.get_or_create(
                teacher=profile,
                subject=form.cleaned_data['subject'],
            )
            messages.success(request, "Assignment added.")
            return redirect('assign_teacher', profile_id=profile.id)
    else:
        form = TeacherAssignmentForm(selected_branch=selected_branch, selected_semester=selected_semester)

    assignments = profile.assignments.select_related('subject').all()
    return render(request, 'attendance/assign_teacher.html', {
        'profile': profile, 'form': form, 'assignments': assignments,
        'selected_branch': selected_branch, 'selected_semester': selected_semester,
    })


@login_required
def remove_assignment(request, assignment_id):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id)
    profile_id = assignment.teacher_id
    assignment.delete()
    return redirect('assign_teacher', profile_id=profile_id)


@login_required
def import_students(request):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            branch = form.cleaned_data['branch']
            semester = int(form.cleaned_data['semester'])
            excel_file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
                created, skipped = 0, 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 3:
                        continue
                    _, usn, name = row[0], row[1], row[2]
                    if not usn or not name:
                        continue
                    usn = str(usn).strip().upper()
                    name = str(name).strip()
                    obj, was_created = Student.objects.get_or_create(
                        usn=usn, branch=branch, semester=semester,
                        defaults={'name': name},
                    )
                    if was_created:
                        created += 1
                    else:
                        skipped += 1
                messages.success(
                    request,
                    f"Import complete for {BRANCH_DICT[branch]} - Semester {semester}: "
                    f"{created} new students added, {skipped} already existed."
                )
                return redirect('admin_dashboard')
            except Exception as exc:
                messages.error(request, f"Could not read the Excel file: {exc}")
    else:
        form = ExcelImportForm()
    return render(request, 'attendance/import_students.html', {'form': form})


@login_required
def admin_branch_detail(request, branch):
    if not is_admin(request.user):
        return HttpResponseForbidden("Admins only.")
    if branch not in BRANCH_DICT:
        return HttpResponseForbidden("Unknown branch.")

    sem_data = []
    for sem, _ in SEMESTER_CHOICES:
        students = Student.objects.filter(branch=branch, semester=sem)
        count = students.count()
        subjects = Subject.objects.filter(branch=branch, semester=sem)
        today_present = Attendance.objects.filter(
            student__branch=branch, student__semester=sem,
            date=timezone.localdate(), status='PRESENT'
        ).count()
        sem_data.append({
            'semester': sem, 'total': count, 'present_today': today_present,
            'subjects': subjects,
        })

    return render(request, 'attendance/admin_branch_detail.html', {
        'branch': branch, 'branch_label': BRANCH_DICT[branch], 'sem_data': sem_data,
    })


@login_required
def admin_chart_data(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'forbidden'}, status=403)

    labels, percentages = [], []
    for code, label in BRANCH_CHOICES:
        total = Attendance.objects.filter(student__branch=code).count()
        present = Attendance.objects.filter(student__branch=code, status='PRESENT').count()
        pct = round((present / total) * 100, 1) if total else 0
        labels.append(code)
        percentages.append(pct)
    return JsonResponse({'labels': labels, 'data': percentages})


# ---------------- TEACHER ----------------

@login_required
def teacher_dashboard(request):
    if request.user.role != 'TEACHER':
        return HttpResponseForbidden("Teachers only.")
    profile = get_object_or_404(TeacherProfile, user=request.user)
    if not profile.is_approved:
        return render(request, 'attendance/teacher_pending.html')

    assignments = profile.assignments.select_related('subject').all()
    return render(request, 'attendance/teacher_dashboard.html', {
        'assignments': assignments, 'branch_dict': BRANCH_DICT,
    })


@login_required
def teacher_attendance_sheet(request, assignment_id):
    """Each subject (assignment) gets its own independent attendance sheet."""
    if request.user.role != 'TEACHER':
        return HttpResponseForbidden("Teachers only.")
    profile = get_object_or_404(TeacherProfile, user=request.user)
    if not profile.is_approved:
        return HttpResponseForbidden("Your account is awaiting admin approval.")

    # Ownership check: a teacher can only open their OWN assignment's sheet.
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id, teacher=profile)
    subject = assignment.subject
    branch, semester = assignment.branch, assignment.semester

    selected_date_str = request.GET.get('date')
    if selected_date_str:
        selected_date = date.fromisoformat(selected_date_str)
    else:
        selected_date = timezone.localdate()

    # Which session of the day (1, 2, 3...) — lets the same subject be
    # marked more than once on the same day (e.g. two periods of Physics).
    try:
        selected_session = int(request.GET.get('session', 1))
        if selected_session < 1:
            selected_session = 1
    except (TypeError, ValueError):
        selected_session = 1

    # Existing sessions already marked for this subject+date, so the
    # dropdown/buttons can show what's already there.
    existing_sessions = sorted(set(
        Attendance.objects.filter(subject=subject, date=selected_date)
        .values_list('session_number', flat=True)
    )) or [1]
    max_existing_session = max(existing_sessions)

    students = Student.objects.filter(branch=branch, semester=semester).order_by('usn')
    today_records = {
        a.student_id: a.status
        for a in Attendance.objects.filter(
            student__in=students, subject=subject, date=selected_date, session_number=selected_session
        )
    }

    rows = []
    for i, student in enumerate(students, start=1):
        # Totals are across ALL sessions/dates for this subject — each
        # session counts as one class, so two periods in a day count as two.
        subject_records = student.attendance_records.filter(subject=subject)
        total_days = subject_records.count()
        present_days = subject_records.filter(status='PRESENT').count()
        absent_days = total_days - present_days
        rows.append({
            'sl_no': i,
            'student': student,
            'status_today': today_records.get(student.id),
            'present_days': present_days,
            'absent_days': absent_days,
            'total_days': total_days,
        })

    context = {
        'assignment': assignment,
        'subject': subject,
        'branch': branch,
        'branch_label': BRANCH_DICT[branch],
        'semester': semester,
        'rows': rows,
        'selected_date': selected_date,
        'selected_session': selected_session,
        'existing_sessions': existing_sessions,
        'next_session': max_existing_session + 1,
        'now': timezone.localtime(),
    }
    return render(request, 'attendance/attendance_sheet.html', context)


@login_required
@require_POST
def mark_attendance(request):
    if request.user.role != 'TEACHER':
        return JsonResponse({'error': 'forbidden'}, status=403)
    profile = get_object_or_404(TeacherProfile, user=request.user)
    if not profile.is_approved:
        return JsonResponse({'error': 'not approved'}, status=403)

    try:
        payload = json.loads(request.body)
        student_id = payload['student_id']
        status = payload['status']
        date_str = payload.get('date')
        assignment_id = payload['assignment_id']
        session_number = int(payload.get('session_number', 1))
    except (KeyError, json.JSONDecodeError, TypeError, ValueError):
        return JsonResponse({'error': 'bad request'}, status=400)

    if status not in ('PRESENT', 'ABSENT'):
        return JsonResponse({'error': 'invalid status'}, status=400)
    if session_number < 1:
        return JsonResponse({'error': 'invalid session number'}, status=400)

    # Ownership check: this assignment (and therefore this subject) must
    # belong to the logged-in teacher — prevents Teacher B from marking
    # Teacher A's subject sheet.
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id, teacher=profile)
    subject = assignment.subject

    student = get_object_or_404(Student, id=student_id)
    if student.branch != assignment.branch or student.semester != assignment.semester:
        return JsonResponse({'error': 'not assigned to this class'}, status=403)

    record_date = date.fromisoformat(date_str) if date_str else timezone.localdate()

    attendance, _ = Attendance.objects.update_or_create(
        student=student, subject=subject, date=record_date, session_number=session_number,
        defaults={'status': status, 'marked_by': request.user},
    )

    subject_records = student.attendance_records.filter(subject=subject)
    total_days = subject_records.count()
    present_days = subject_records.filter(status='PRESENT').count()
    absent_days = total_days - present_days

    return JsonResponse({
        'ok': True,
        'status': attendance.status,
        'present_days': present_days,
        'absent_days': absent_days,
        'total_days': total_days,
        'marked_at': timezone.localtime(attendance.marked_at).strftime('%d-%b-%Y %I:%M %p'),
    })


@login_required
def teacher_chart_data(request, assignment_id):
    if request.user.role != 'TEACHER':
        return JsonResponse({'error': 'forbidden'}, status=403)
    profile = get_object_or_404(TeacherProfile, user=request.user)
    assignment = get_object_or_404(TeacherAssignment, id=assignment_id, teacher=profile)
    subject = assignment.subject

    students = Student.objects.filter(branch=assignment.branch, semester=assignment.semester)
    labels, present_counts, absent_counts = [], [], []
    for s in students:
        subject_records = s.attendance_records.filter(subject=subject)
        total = subject_records.count()
        present = subject_records.filter(status='PRESENT').count()
        labels.append(s.usn)
        present_counts.append(present)
        absent_counts.append(total - present)
    return JsonResponse({'labels': labels, 'present': present_counts, 'absent': absent_counts})


# ---------------- STUDENT (read-only) ----------------

@login_required
def student_dashboard(request):
    if request.user.role != 'STUDENT':
        return HttpResponseForbidden("Students only.")
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return render(request, 'attendance/student_no_record.html')

    all_records = student.attendance_records.all().order_by('-date')
    total_days = all_records.count()
    present_days = all_records.filter(status='PRESENT').count()
    absent_days = total_days - present_days
    percentage = round((present_days / total_days) * 100, 1) if total_days else 0

    # Group attendance by subject so the student sees per-subject percentages,
    # not just one blended overall number.
    subjects_qs = Subject.objects.filter(branch=student.branch, semester=student.semester)
    subject_breakdown = []
    for subject in subjects_qs:
        subj_records = all_records.filter(subject=subject)
        subj_total = subj_records.count()
        subj_present = subj_records.filter(status='PRESENT').count()
        subj_absent = subj_total - subj_present
        subj_pct = round((subj_present / subj_total) * 100, 1) if subj_total else 0
        subject_breakdown.append({
            'subject': subject,
            'total_days': subj_total,
            'present_days': subj_present,
            'absent_days': subj_absent,
            'percentage': subj_pct,
            'records': subj_records[:60],
        })

    return render(request, 'attendance/student_dashboard.html', {
        'student': student,
        'records': all_records[:60],
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'percentage': percentage,
        'branch_label': BRANCH_DICT[student.branch],
        'subject_breakdown': subject_breakdown,
    })


# ---------------- ABOUT PAGE ----------------

def about_page(request):
    """Public 'About this app' page — no login required, so it also works
    as a nice landing/credits page."""
    return render(request, 'attendance/about.html')
